"""
ContactIQ — File Manager
Handles file uploads, validation, template generation, and data source switching.
Maps the 8 data files (5 raw + 3 config) to upload categories.
"""
import os
import shutil
import json
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads', 'active')
TEMPLATE_DIR = os.path.join(BASE_DIR, 'uploads', 'templates')

# ── File Categories ──────────────────────────────────────────
# Maps each uploadable file to its category, default path, and expected schema.

FILE_REGISTRY = {
    'ccaas_interactions': {
        'label': 'CCaaS Interaction Records',
        'description': 'Contact center interaction data — volumes, AHT, CSAT, channels, queues',
        'default_file': 'raw/ccaas_interaction_records.xlsx',
        'required': True,
        'expected_columns': [
            'Queue Name', 'Interval Start', 'Contacts Offered', 'Contacts Handled',
            'Avg Handle Time', 'Avg Speed Answer', 'Abandoned', 'Service Level',
            'Avg After Call Work'
        ],
        'icon': '📞',
    },
    'crm_cases': {
        'label': 'CRM Case Export',
        'description': 'CRM case records — case types, resolution, customer details',
        'default_file': 'raw/crm_case_export.xlsx',
        'required': True,
        'expected_columns': [
            'Case ID', 'Created Date', 'Case Type', 'Sub Type', 'Status',
            'Channel', 'Resolution', 'Customer Satisfaction'
        ],
        'icon': '📋',
    },
    'hr_workforce': {
        'label': 'HR Workforce Data',
        'description': 'Workforce roster — headcounts, roles, salaries, locations',
        'default_file': 'raw/hr_workforce_data.xlsx',
        'required': True,
        'expected_columns': [
            'Employee ID', 'Job Title', 'Department', 'Location', 'Start Date',
            'Annual Salary', 'Employment Type', 'Status'
        ],
        'icon': '👥',
    },
    'interaction_transcripts': {
        'label': 'Interaction Transcripts',
        'description': 'Call/chat transcripts — for intent analysis and NLP scoring',
        'default_file': 'raw/interaction_transcripts.xlsx',
        'required': False,
        'expected_columns': [
            'Interaction ID', 'Queue', 'Channel', 'Transcript Text', 'Duration',
            'Agent ID', 'Timestamp'
        ],
        'icon': '💬',
    },
    'survey_responses': {
        'label': 'Survey Responses',
        'description': 'Customer satisfaction surveys — CSAT, NPS, CES scores',
        'default_file': 'raw/survey_responses.xlsx',
        'required': False,
        'expected_columns': [
            'Response ID', 'Survey Date', 'Channel', 'Queue', 'CSAT Score',
            'NPS Score', 'CES Score', 'Comments'
        ],
        'icon': '⭐',
    },
    'parameters': {
        'label': 'Parameters & Configuration',
        'description': 'Client parameters — financials, planning horizon, thresholds',
        'default_file': 'config/parameters.xlsx',
        'required': True,
        'expected_columns': ['Parameter', 'Value', 'Description'],
        'icon': '⚙️',
    },
    'benchmarks': {
        'label': 'Industry Benchmarks',
        'description': 'Benchmark data — industry averages, top/bottom quartile by metric',
        'default_file': 'config/benchmarks.xlsx',
        'required': False,
        'expected_columns': [
            'Scope', 'Intent', 'Channel', 'Metric', 'Industry Average',
            'Top Quartile', 'Bottom Quartile', 'Source'
        ],
        'icon': '📊',
    },
    'technology_investment': {
        'label': 'Technology & Investment',
        'description': 'Investment assumptions — technology costs, implementation timelines',
        'default_file': 'config/technology_investment.xlsx',
        'required': False,
        'expected_columns': [
            'Initiative ID', 'Technology', 'One-Time Cost', 'Annual Recurring',
            'Implementation Months', 'Vendor', 'Category'
        ],
        'icon': '💰',
    },
}


def ensure_dirs():
    """Create upload and template directories if they don't exist."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(TEMPLATE_DIR, exist_ok=True)


def get_file_status():
    """
    Get status of all file categories.
    Returns dict: category -> {label, status: 'backend'|'uploaded'|'missing', ...}
    """
    ensure_dirs()
    status = {}
    for cat, info in FILE_REGISTRY.items():
        default_path = os.path.join(DATA_DIR, info['default_file'])
        upload_path = os.path.join(UPLOAD_DIR, f"{cat}.xlsx")

        if os.path.exists(upload_path):
            stat = os.stat(upload_path)
            status[cat] = {
                **info,
                'status': 'uploaded',
                'active_path': upload_path,
                'file_size': stat.st_size,
                'uploaded_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            }
        elif os.path.exists(default_path):
            status[cat] = {
                **info,
                'status': 'backend',
                'active_path': default_path,
                'file_size': os.path.getsize(default_path),
            }
        else:
            status[cat] = {
                **info,
                'status': 'missing',
                'active_path': None,
            }
    return status


def get_active_file_path(category):
    """
    Get the active file path for a category.
    Returns uploaded file if exists, otherwise backend default.
    """
    ensure_dirs()
    if category not in FILE_REGISTRY:
        return None

    upload_path = os.path.join(UPLOAD_DIR, f"{category}.xlsx")
    if os.path.exists(upload_path):
        return upload_path

    default_path = os.path.join(DATA_DIR, FILE_REGISTRY[category]['default_file'])
    if os.path.exists(default_path):
        return default_path

    return None


def get_data_dir_for_engine():
    """
    Get the effective data directory paths for the engine.
    If uploads exist, creates a merged view with uploads overriding defaults.
    Returns a dict mapping default filenames to their active paths.
    """
    ensure_dirs()
    path_map = {}
    for cat, info in FILE_REGISTRY.items():
        active = get_active_file_path(cat)
        if active:
            path_map[cat] = active
            # Also map by default filename for backward compatibility
            path_map[info['default_file']] = active
    return path_map


def validate_upload(filepath, category):
    """
    Validate an uploaded file against expected schema.
    Returns (is_valid, notes).
    """
    if category not in FILE_REGISTRY:
        return False, f"Unknown category: {category}"

    expected = FILE_REGISTRY[category].get('expected_columns', [])
    if not expected:
        return True, "No schema validation required"

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True, max_row=1))
        wb.close()

        if not rows:
            return False, "File is empty — no header row found"

        headers = [str(h).strip() if h else '' for h in rows[0]]
        missing = [col for col in expected if col not in headers]

        if not missing:
            return True, f"Schema validated — all {len(expected)} expected columns found"

        # Partial match — warn but allow
        found = len(expected) - len(missing)
        pct = round(found / len(expected) * 100)
        if pct >= 60:
            return True, f"Partial match ({pct}%): missing columns: {', '.join(missing[:5])}"
        else:
            return False, f"Poor match ({pct}%): missing columns: {', '.join(missing[:5])}"

    except Exception as e:
        return False, f"Validation error: {str(e)}"


def save_uploaded_file(file_storage, category):
    """
    Save an uploaded file for a category.
    file_storage: werkzeug FileStorage object from request.files.
    Returns (success, message, filepath).
    """
    ensure_dirs()
    if category not in FILE_REGISTRY:
        return False, f"Unknown category: {category}", None

    # Save to a temp location first for validation
    filename = f"{category}.xlsx"
    filepath = os.path.join(UPLOAD_DIR, filename)
    temp_path = filepath + '.tmp'

    try:
        file_storage.save(temp_path)
        file_size = os.path.getsize(temp_path)

        if file_size == 0:
            os.remove(temp_path)
            return False, "Uploaded file is empty", None

        if file_size > 50 * 1024 * 1024:  # 50MB limit
            os.remove(temp_path)
            return False, "File exceeds 50MB limit", None

        # Validate schema
        is_valid, notes = validate_upload(temp_path, category)

        if not is_valid:
            os.remove(temp_path)
            return False, notes, None

        # Move temp to active
        shutil.move(temp_path, filepath)
        return True, notes, filepath

    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return False, f"Upload failed: {str(e)}", None


def clear_uploaded_file(category):
    """
    Remove uploaded override for a category (revert to backend data).
    Returns (success, message).
    """
    ensure_dirs()
    if category not in FILE_REGISTRY:
        return False, f"Unknown category: {category}"

    filepath = os.path.join(UPLOAD_DIR, f"{category}.xlsx")
    if os.path.exists(filepath):
        os.remove(filepath)
        return True, f"Upload cleared for {FILE_REGISTRY[category]['label']} — reverted to backend data"
    return True, f"No upload to clear for {FILE_REGISTRY[category]['label']}"


def clear_all_uploads():
    """Remove all uploaded files, revert everything to backend."""
    ensure_dirs()
    cleared = 0
    for cat in FILE_REGISTRY:
        filepath = os.path.join(UPLOAD_DIR, f"{cat}.xlsx")
        if os.path.exists(filepath):
            os.remove(filepath)
            cleared += 1
    return cleared


def generate_template(category):
    """
    Generate a downloadable template XLSX for a category.
    Returns the template file path.
    """
    ensure_dirs()
    if category not in FILE_REGISTRY:
        return None

    info = FILE_REGISTRY[category]
    template_path = os.path.join(TEMPLATE_DIR, f"{category}_template.xlsx")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = info['label'][:31]  # Excel sheet name max 31 chars

        # Header styling
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='2E2E38', end_color='2E2E38', fill_type='solid')
        tb = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        columns = info.get('expected_columns', [])
        for c, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = tb
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 15)

        # Add a sample row with placeholder text
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=2, column=c, value='[Sample Data]')
            cell.border = tb
            cell.font = Font(color='999999', italic=True)

        # Add instructions sheet
        ws2 = wb.create_sheet('Instructions')
        ws2['A1'] = f"ContactIQ — {info['label']} Template"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A3'] = info['description']
        ws2['A5'] = 'Required Columns:'
        ws2['A5'].font = Font(bold=True)
        for i, col in enumerate(columns):
            ws2.cell(row=6 + i, column=1, value=f"  • {col}")
        ws2.column_dimensions['A'].width = 60

        wb.save(template_path)
        return template_path

    except ImportError:
        return None
    except Exception:
        return None


def get_upload_summary():
    """Get a concise summary of current data state for the UI."""
    status = get_file_status()
    summary = {
        'total_files': len(FILE_REGISTRY),
        'uploaded': sum(1 for s in status.values() if s['status'] == 'uploaded'),
        'backend': sum(1 for s in status.values() if s['status'] == 'backend'),
        'missing': sum(1 for s in status.values() if s['status'] == 'missing'),
        'categories': status,
    }
    summary['data_mode'] = 'upload' if summary['uploaded'] > 0 else 'backend'
    return summary
