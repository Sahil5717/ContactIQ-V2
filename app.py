"""
ContactIQ — Intelligent Contact Center Optimization Platform (v8 Infrastructure)
Phase 1 Infrastructure: Authentication, File Upload, Data Management, API Layer.
All existing engine routes preserved. New: login, file management, data source switching.
"""
import json
import os
import tempfile
import copy
import traceback
import threading
from flask import Flask, jsonify, request, render_template, send_file, session, redirect, g
from engines.data_loader import run_etl, set_path_overrides
from engines.diagnostic import run_diagnostic
from engines.maturity import run_maturity
from engines.readiness import compute_readiness, STRATEGIC_DRIVERS
from engines.waterfall import score_initiatives, run_waterfall, compute_pillar_scenarios
from engines.risk import run_risk
from engines.workforce import run_workforce
from engines.channel_strategy import run_channel_strategy
from engines.transcripts import run_transcript_analysis
from engines.insights import run_insights
from engines.scenarios import compare_scenarios, compute_delta
from engines.recommendations import get_recommendations, get_initiative_linkage, get_available_industries, get_industry_config
from infrastructure.database import init_db, validate_session, load_overrides, save_overrides, get_project_mode, set_project_mode, get_mode_change_log
from infrastructure.auth import init_auth, login_user, logout_user, get_current_user, require_role
from infrastructure.file_manager import (
    FILE_REGISTRY, get_file_status, get_active_file_path, get_upload_summary,
    save_uploaded_file, clear_uploaded_file, clear_all_uploads, generate_template,
    UPLOAD_DIR, ensure_dirs
)

app = Flask(__name__)
# A-07 fix: Require SECRET_KEY in production, generate random key for dev only
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    import secrets
    _secret = secrets.token_hex(32)
    print("[WARN] SECRET_KEY not set — using random key (sessions won't survive restart)")
app.config['SECRET_KEY'] = _secret
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB upload limit

STATE = {
    'data': None, 'diagnostic': None, 'maturity': None,
    'readiness': None, 'initiatives': None, 'waterfall': None,
    'pillarScenarios': None,
    'risk': None, 'workforce': None, 'channelStrategy': None,
    'transcriptAnalysis': None, 'insights': None, 'scenarioComparison': None,
    'overrides': {}, 'loaded': False,
}

# A-08 fix: Lock for STATE mutations to prevent concurrent request interleaving
STATE_LOCK = threading.Lock()

# ── Initialize Infrastructure ────────────────────────────────
init_db()
init_auth(app)

# ── V9: Security Headers ──
@app.after_request
def _set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# ── V9: Session cookie security ──
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# ── V9: Login rate limiting ──
_login_attempts = {}  # ip -> (count, first_attempt_time)
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SEC = 900  # 15 minutes

def _check_rate_limit(ip):
    """Check if IP is rate limited. Returns (allowed, retry_after_sec)."""
    import time
    now = time.time()
    if ip in _login_attempts:
        count, first_time = _login_attempts[ip]
        if now - first_time > _LOGIN_WINDOW_SEC:
            _login_attempts[ip] = (1, now)
            return True, 0
        if count >= _LOGIN_MAX_ATTEMPTS:
            return False, int(_LOGIN_WINDOW_SEC - (now - first_time))
        _login_attempts[ip] = (count + 1, first_time)
        return True, 0
    _login_attempts[ip] = (1, now)
    return True, 0

# A-02 fix: Session check for all API data routes
@app.before_request
def _check_api_auth():
    """Require a valid session for all /api/ routes except health and auth."""
    path = request.path
    if path.startswith('/api/') and not path.startswith('/api/auth/') and path != '/api/health':
        user = get_current_user()
        if not user:
            # Allow unauthenticated access in dev when no users exist
            from infrastructure.database import get_db
            try:
                with get_db() as db:
                    count = db.execute("SELECT COUNT(*) as c FROM users").fetchone()['c']
                    if count == 0:
                        return  # No users configured — allow access (initial setup)
            except Exception:
                # F-03 fix: DB failure must deny access, not grant it
                return jsonify({'error': 'Service temporarily unavailable'}), 503
            return jsonify({'error': 'Authentication required'}), 401


# F-18 fix: Periodically clean up expired sessions (1 in 100 requests)
import random
@app.before_request
def _maybe_cleanup_sessions():
    if random.randint(1, 100) == 1:
        try:
            from infrastructure.database import cleanup_expired_sessions
            cleanup_expired_sessions()
        except Exception:
            pass  # Non-critical — don't block the request


def _sanitize_for_json(obj):
    """Recursively sanitize objects for JSON serialization."""
    if isinstance(obj, set): return sorted(list(obj))
    elif isinstance(obj, dict): return {k: _sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list): return [_sanitize_for_json(v) for v in obj]
    return obj


def _apply_file_overrides():
    """Check for uploaded files and set path overrides before running ETL."""
    ensure_dirs()
    overrides = {}
    for cat in FILE_REGISTRY:
        upload_path = os.path.join(UPLOAD_DIR, f"{cat}.xlsx")
        if os.path.exists(upload_path):
            overrides[cat] = upload_path
    set_path_overrides(overrides)


def _run_all():
    """CR-FIX-AH: Pipeline refactored into named stages for testability and maintainability."""
    with STATE_LOCK:
        _apply_file_overrides()
        
        # ── Stage 1: Data Loading ──
        _stage_load_data()
        
        # ── Stage 2: Diagnostic & Readiness ──
        _stage_diagnostic()
        
        # ── Stage 3: Initiative Scoring ──
        _stage_initiatives()
        
        # ── Stage 4: Waterfall & Financial Projection ──
        _stage_waterfall()
        
        # ── Stage 5: Downstream Engines ──
        _stage_downstream()
        
        # ── Stage 6: Intelligence & Enrichment ──
        _stage_enrichment()
        
        STATE['loaded'] = True
        return True


def _stage_load_data():
    """Stage 1: ETL pipeline — load and validate all source data."""
    data = run_etl()
    if 'totalCost' not in data:
        data['totalCost'] = sum(r['headcount'] * r['costPerFTE'] for r in data['roles'])
    if 'avgCPC' not in data:
        annual_vol = data.get('totalVolumeAnnual', data.get('totalVolume', 1))
        data['avgCPC'] = round(data['totalCost'] / max(annual_vol, 1), 2)
    STATE['data'] = data


def _stage_diagnostic():
    """Stage 2: Run diagnostic, maturity, and readiness assessments."""
    data = STATE['data']
    STATE['diagnostic'] = run_diagnostic(data)
    STATE['maturity'] = run_maturity(data, STATE['diagnostic'])
    STATE['readiness'] = compute_readiness(data, STATE['diagnostic'], STATE['maturity'])


def _stage_initiatives():
    """Stage 3: Score and select initiatives, apply persisted overrides."""
    data = STATE['data']
    STATE['initiatives'] = score_initiatives(data, STATE['diagnostic'], STATE['readiness'])
    persisted = load_overrides()
    if persisted:
        STATE['overrides'].update(persisted)
    _apply_all_overrides()


def _stage_waterfall():
    """Stage 4: Run waterfall cascade and pillar scenarios."""
    data = STATE['data']
    STATE['waterfall'] = run_waterfall(data, STATE['initiatives'])
    try:
        STATE['pillarScenarios'] = compute_pillar_scenarios(data, STATE['initiatives'])
    except Exception as e:
        print(f"[WARN] Pillar scenarios failed: {e}")
        STATE['pillarScenarios'] = None


def _stage_downstream():
    """Stage 5: Risk, workforce, and channel strategy engines."""
    data = STATE['data']
    STATE['risk'] = run_risk(STATE['initiatives'], data)
    STATE['workforce'] = run_workforce(data, STATE['waterfall'], STATE['initiatives'])
    STATE['channelStrategy'] = run_channel_strategy(data, STATE['diagnostic'], STATE['initiatives'])
    STATE['subIntentAnalysis'] = STATE['diagnostic'].get('subIntentAnalysis', [])


def _stage_enrichment():
    """Stage 6: Transcript analysis, insights, scenario comparison."""
    data = STATE['data']
    try:
        STATE['transcriptAnalysis'] = run_transcript_analysis()
    except Exception as e:
        print(f"[WARN] Transcript analysis failed: {e}")
        STATE['transcriptAnalysis'] = {}
    try:
        STATE['insights'] = run_insights(data, STATE['diagnostic'], STATE['waterfall'])
    except Exception as e:
        print(f"[WARN] Insight engine failed: {e}")
        STATE['insights'] = {}
    try:
        STATE['scenarioComparison'] = compare_scenarios(data, STATE['initiatives'], run_waterfall)
    except Exception as e:
        print(f"[WARN] Scenario comparison failed: {e}")
        STATE['scenarioComparison'] = []


def _recompute_downstream():
    with STATE_LOCK:
        data = STATE['data']
        STATE['waterfall'] = run_waterfall(data, STATE['initiatives'])
        try:
            STATE['pillarScenarios'] = compute_pillar_scenarios(data, STATE['initiatives'])
        except Exception:
            pass
        STATE['risk'] = run_risk(STATE['initiatives'], data)
        STATE['workforce'] = run_workforce(data, STATE['waterfall'], STATE['initiatives'])


def _recompute_all_from_diagnostic():
    with STATE_LOCK:
        _recompute_all_from_diagnostic_unlocked()


def _recompute_all_from_diagnostic_unlocked():
    """Internal helper — called from within STATE_LOCK, no re-acquire."""
    data = STATE['data']
    STATE['diagnostic'] = run_diagnostic(data)
    STATE['maturity'] = run_maturity(data, STATE['diagnostic'])
    STATE['readiness'] = compute_readiness(data, STATE['diagnostic'], STATE['maturity'])
    STATE['initiatives'] = score_initiatives(data, STATE['diagnostic'], STATE['readiness'])
    _apply_all_overrides()
    _recompute_downstream_unlocked()
    STATE['channelStrategy'] = run_channel_strategy(data, STATE['diagnostic'], STATE['initiatives'])


def _recompute_downstream_unlocked():
    """Internal helper — called from within STATE_LOCK, no re-acquire."""
    data = STATE['data']
    # V9: Snapshot previous state for delta view
    prev_wf = STATE.get('waterfall') or {}
    prev_snapshot = {
        'totalReduction': prev_wf.get('totalReduction', 0),
        'totalSaving': prev_wf.get('totalSaving', 0),
        'totalNPV': prev_wf.get('totalNPV', 0),
        'totalInvestment': prev_wf.get('totalInvestment', 0),
        'enabledCount': len([i for i in STATE.get('initiatives', []) if i.get('enabled')]),
    }
    
    STATE['waterfall'] = run_waterfall(data, STATE['initiatives'])
    try:
        STATE['pillarScenarios'] = compute_pillar_scenarios(data, STATE['initiatives'])
    except Exception:
        pass
    STATE['risk'] = run_risk(STATE['initiatives'], data)
    STATE['workforce'] = run_workforce(data, STATE['waterfall'], STATE['initiatives'])
    
    # V9: Compute delta
    new_wf = STATE['waterfall']
    STATE['lastDelta'] = {
        'fteChange': new_wf.get('totalReduction', 0) - prev_snapshot['totalReduction'],
        'savingChange': new_wf.get('totalSaving', 0) - prev_snapshot['totalSaving'],
        'npvChange': new_wf.get('totalNPV', 0) - prev_snapshot['totalNPV'],
        'investmentChange': new_wf.get('totalInvestment', 0) - prev_snapshot['totalInvestment'],
        'previous': prev_snapshot,
        'current': {
            'totalReduction': new_wf.get('totalReduction', 0),
            'totalSaving': new_wf.get('totalSaving', 0),
            'totalNPV': new_wf.get('totalNPV', 0),
            'totalInvestment': new_wf.get('totalInvestment', 0),
            'enabledCount': len([i for i in STATE['initiatives'] if i.get('enabled')]),
        },
    }


def _apply_all_overrides():
    # F-27 WARNING: This function is called from within STATE_LOCK (by _run_all,
    # _recompute_all_from_diagnostic_unlocked, api_reset_pillar_toggles).
    # It MUST remain lock-free — never acquire STATE_LOCK or call any function
    # that acquires STATE_LOCK from here, or it will deadlock.
    # Initiative overrides
    for init in STATE['initiatives']:
        iid = init['id']
        ok = f"init_enabled_{iid}"
        if ok in STATE['overrides']: init['enabled'] = STATE['overrides'][ok]
        for rk in ('rampYear1','rampYear2','rampYear3'):
            rkey = f"init_{rk}_{iid}"
            if rkey in STATE['overrides']: init[rk] = STATE['overrides'][rkey]
        field_key = f"init_fields_{iid}"
        if field_key in STATE['overrides']:
            for fk, fv in STATE['overrides'][field_key].items():
                init[fk] = fv
    # v12-#53: Benchmark overrides — re-apply saved benchmark values after recalculation
    for key, val in STATE['overrides'].items():
        if key.startswith('benchmark_'):
            metric = key.replace('benchmark_', '')
            bm = STATE['data'].get('benchmarks', {})
            if metric in bm:
                if isinstance(bm[metric], dict):
                    bm[metric]['global'] = val
                else:
                    bm[metric] = val


def _build_demo_object(overrides=None):
    ov = overrides or {}
    data = STATE['data']; diag = STATE['diagnostic']; mat = STATE['maturity']
    wf = ov.get('waterfall', STATE['waterfall'])
    rsk = ov.get('risk', STATE['risk'])
    wkf = ov.get('workforce', STATE['workforce'])
    chs = STATE['channelStrategy']; roles = data['roles']
    inits = ov.get('initiatives', STATE['initiatives'])
    return {
        'queues': data['queues'], 'roles': roles, 'params': data['params'],
        'benchmarks': data.get('benchmarks', {}),
        'totalVolume': data['totalVolume'], 'totalFTE': data['totalFTE'],
        'totalCost': data['totalCost'],
        'totalVolumeAnnual': data.get('totalVolumeAnnual', data['totalVolume']),
        'annualContactsPerFTE': round(data.get('totalVolumeAnnual', data['totalVolume']) / max(data['totalFTE'], 1)),
        'volumeAnnualizationFactor': data.get('volumeAnnualizationFactor', 12),
        'totalMonthlyCost': sum(r['headcount']*r['costPerFTE']/12 for r in roles),
        'avgCSAT': data['avgCSAT'], 'avgAHT': round(data['avgAHT']*60, 0),
        'avgAHT_min': round(data['avgAHT'], 1),
        'avgFCR': data.get('avgFCR', 0), 'avgCPC': data.get('avgCPC', 0),
        'avgEscalation': data.get('avgEscalation', _weighted_avg(data['queues'], 'escalation')),
        'avgRepeat': data.get('avgRepeat', _weighted_avg(data['queues'], 'repeat')),
        'avgCES': data.get('avgCES', _weighted_avg(data['queues'], 'ces')),
        'channelMix': _build_channel_mix(data['queues']),
        'buMix': _build_bu_mix(data['queues']),
        'intentMix': _build_intent_mix(data['queues']),
        'healthScores': diag.get('queueScores', []),
        'overallHealth': diag.get('summary', {}).get('avgScore', 0),
        'healthRag': 'green' if diag.get('summary',{}).get('avgScore',0) >= 70 else 'amber' if diag.get('summary',{}).get('avgScore',0) >= 40 else 'red',
        'problemAreas': diag.get('problemAreas', []),
        'rootCauses': diag.get('rootCauses', []),
        'costAnalysis': diag.get('costAnalysis', {}),
        'channelSummary': diag.get('channelSummary', []),
        'subIntentAnalysis': diag.get('subIntentAnalysis', []),
        'mismatch': diag.get('mismatch', []),
        'mismatchSummary': diag.get('mismatch', []),
        'mismatchDetail': diag.get('mismatch', []),
        'maturityScores': mat.get('dimensions', {}),
        'maturityOverall': mat.get('overall', 0),
        'maturityLevel': mat.get('overallLevel', 1),
        'maturityLabel': mat.get('levelInfo', {}).get('label', ''),
        'maturityGaps': mat.get('gaps', []),
        'maturityRadar': mat.get('radar', {}),
        'maturityRecommendations': mat.get('gaps', []),
        'initiatives': inits,
        'enabledCount': sum(1 for i in inits if i.get('enabled')),
        'waterfall': wf,
        'yearlyProjections': wf.get('yearly', []),
        'scenarios': wf.get('scenarios', {}),
        'sensitivity': wf.get('sensitivity', []),
        'roleImpact': wf.get('roleImpact', {}),
        'investmentItems': wf.get('investmentItems', []),
        'investmentYearly': wf.get('investmentYearly', []),
        'investmentSummary': wf.get('investmentSummary', {}),
        'financials': {
            'totalNPV': wf.get('totalNPV',0), 'totalSaving': wf.get('totalSaving',0),
            'totalInvestment': wf.get('totalInvestment',0), 'roi': wf.get('roi',0),
            'roiGross': wf.get('roiGross',0), 'payback': wf.get('payback',0),
            'irr': wf.get('irr',0), 'techInvestment': wf.get('techInvestment',0),
            'annualMaintenance': wf.get('annualMaintenance',0),
        },
        'riskRegister': rsk.get('initiatives', []),
        'riskSummary': rsk.get('summary', {}),
        'topRisks': sorted(rsk.get('initiatives',[]), key=lambda x: x.get('overallRisk',0), reverse=True)[:5],
        'riskByLayer': rsk.get('riskByLayer', {}),
        'riskByBU': rsk.get('riskByBU', {}),
        'riskDependencies': rsk.get('dependencies', {}),
        'workforceTransition': wkf.get('transitions', []),
        'workforceSummary': wkf.get('summary', {}),
        'reskillMatrix': wkf.get('reskillMatrix', {}),
        'reskillDemand': wkf.get('reskillDemand', {}),
        'workforceByLocation': wkf.get('byLocation', {}),
        'workforceByBU': wkf.get('byBU', {}),
        'heatmapData': _build_heatmap(data['queues']),
        'costBreakdown': _build_cost_breakdown(data),
        'channelRecommendations': chs.get('recommendations', []),
        'channelSankey': chs.get('sankey', {}),
        'currentDigitalPct': chs.get('currentDigitalPct', 0),
        'targetDigitalPct': chs.get('targetDigitalPct', 0),
        'channelScorecard': chs.get('recommendations', []),
        'channelIntroductions': [r for r in chs.get('recommendations',[]) if r.get('decision')=='invest'],
        'channelRetirements': [r for r in chs.get('recommendations',[]) if r.get('decision') in ('sunset','migrate_from')],
        'channelOptimalMix': {'currentDigital': chs.get('currentDigitalPct',0), 'targetDigital': chs.get('targetDigitalPct',0)},
        'intentMatrix': chs.get('intentMatrix', []),
        'targetMix': chs.get('targetMix', {}),
        'channelMigrations': chs.get('migrations', []),
        'migrationReadiness': chs.get('migrationReadiness', {}),
        'frictionSignals': chs.get('frictionSignals', []),
        'cxSafeguards': chs.get('cxSafeguards', {}),
        'channelCostAnalysis': chs.get('costAnalysis', {}),
        'migrationSavings': chs.get('migrationSavings', {}),
        'readiness': _sanitize_for_json(STATE.get('readiness', {})),
        'automationReadiness': STATE.get('readiness', {}).get('automationReadiness', 0),
        'opModelGap': STATE.get('readiness', {}).get('opModelGap', 0),
        'locationScore': STATE.get('readiness', {}).get('locationScore', 0),
        'strategicDriver': STATE.get('readiness', {}).get('strategicDriver', 'cost_optimization'),
        'strategicDrivers': STRATEGIC_DRIVERS,
        'layerFTE': wf.get('layerFTE', {}),
        'layerSaving': wf.get('layerSaving', {}),
        'poolUtilization': wf.get('poolUtilization', {}),
        'poolSummary': wf.get('poolSummary', {}),
        'auditTrail': wf.get('auditTrail', []),
        'clientName': data['params'].get('clientName','Client'),
        'industry': data['params'].get('industry','Custom'),
        'currency': data['params'].get('currency','USD'),
        'horizon': data['params'].get('horizon',3),
        # P2-1: Dimensional engine fields
        'locations': data.get('locations', ['Onshore']),
        'sourcingTypes': data.get('sourcingTypes', ['In-house']),
        'locationCostMatrix': data.get('locationCostMatrix', {}),
        'buSummary': wf.get('buSummary', {}),
        'bus': data.get('bus', []),
        # v12-#35: Sub-intent data for downstream pages
        'subIntentEnriched': _enrich_sub_intents_for_downstream(
            diag.get('subIntentAnalysis', []), inits, wf),
        # CR-FIX-VOL: Volume scaling transparency
        'volumeScaling': data.get('volumeScaling', {}),
        # CR-FIX-TAG: Metric provenance tagging
        'metricSources': data.get('metricSources', {}),
        # CR-FIX-WFM: WFM actuals (shrinkage decomposition, occupancy, utilization)
        'wfmActuals': data['params'].get('_wfmActuals', {}),
        # CR-FIX-CONF: Confidence bands and data quality
        'confidenceBands': wf.get('confidenceBands', {}),
        'dataQuality': wf.get('dataQuality', {}),
        # CR-FIX-TRANSCRIPT: Transcript analysis
        'transcriptAnalysis': _sanitize_for_json(STATE.get('transcriptAnalysis', {})),
        # CR-FIX-BASIS: Calculation basis and validation
        'calculationBasis': data.get('calculationBasis', {}),
        'validationIssues': data.get('validationIssues', []),
        # CR-FIX-O: Insights
        'insights': _sanitize_for_json(STATE.get('insights', {})),
        # CR-FIX-R: Scenario comparison
        'scenarioComparison': STATE.get('scenarioComparison', []),
        # CR-FIX-AG: Model version
        'modelVersion': '8.0.0',
    }


# ══════════════════════════════════════════════════════════════
#  AUTHENTICATION ROUTES
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
#  HEALTH CHECK (Railway deployment)
# ══════════════════════════════════════════════════════════════

@app.route('/api/health')
def api_health():
    return jsonify({'status': 'ok', 'version': 'v8.0', 'service': 'ContactIQ'})


@app.route('/api/audit-trail')
def api_audit_trail():
    """CR-FIX-AUDIT: Return override audit trail."""
    from infrastructure.database import get_override_audit_log
    try:
        log = get_override_audit_log(limit=200)
        return jsonify({'auditLog': log, 'count': len(log)})
    except Exception as e:
        return jsonify({'auditLog': [], 'count': 0, 'error': str(e)})


@app.route('/login')
def login_page():
    if session.get('auth_token') and validate_session(session['auth_token']):
        return redirect('/')
    return render_template('login.html')


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    # V9: Rate limiting
    ip = request.remote_addr or '0.0.0.0'
    allowed, retry_after = _check_rate_limit(ip)
    if not allowed:
        return jsonify({'error': f'Too many login attempts. Try again in {retry_after} seconds.'}), 429
    body = request.get_json(force=True)
    username = body.get('username', '').strip()
    password = body.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    success, result, token = login_user(username, password, ip_address=request.remote_addr)
    if not success:
        return jsonify({'error': result}), 401
    session['auth_token'] = token
    return jsonify({
        'status': 'ok',
        'user': {'id': result['id'], 'username': result['username'],
                 'role': result['role'], 'display_name': result['display_name']},
        'redirect': '/',
    })


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    logout_user()
    return jsonify({'status': 'ok', 'redirect': '/login'})


@app.route('/api/auth/me')
def api_auth_me():
    user = get_current_user()
    if not user: return jsonify({'error': 'Not authenticated'}), 401
    return jsonify({'id': user['id'], 'username': user['username'],
                    'role': user['role'], 'display_name': user['display_name']})


# ══════════════════════════════════════════════════════════════
#  DATA MANAGEMENT ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/api/data-management/status')
def api_data_status():
    return jsonify(get_upload_summary())


@app.route('/api/data-management/upload', methods=['POST'])
@require_role('analyst')
def api_upload_file():
    category = request.form.get('category')
    if not category or category not in FILE_REGISTRY:
        return jsonify({'error': f'Valid category required'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx files accepted'}), 400
    success, message, filepath = save_uploaded_file(file, category)
    if not success:
        return jsonify({'error': message}), 400
    return jsonify({'status': 'ok', 'message': message, 'category': category,
                    'label': FILE_REGISTRY[category]['label']})


@app.route('/api/data-management/clear', methods=['POST'])
@require_role('supervisor')
def api_clear_upload():
    body = request.get_json(force=True)
    category = body.get('category')
    if category == 'all':
        cleared = clear_all_uploads()
        return jsonify({'status': 'ok', 'message': f'Cleared {cleared} uploads'})
    if not category or category not in FILE_REGISTRY:
        return jsonify({'error': 'Valid category required'}), 400
    success, message = clear_uploaded_file(category)
    return jsonify({'status': 'ok', 'message': message})


@app.route('/api/data-management/template/<category>')
def api_download_template(category):
    if category not in FILE_REGISTRY:
        return jsonify({'error': f'Unknown category: {category}'}), 404
    template_path = generate_template(category)
    if not template_path:
        return jsonify({'error': 'Could not generate template'}), 500
    return send_file(template_path, as_attachment=True,
                     download_name=f'{category}_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/data-management/recalculate', methods=['POST'])
@require_role('supervisor')
def api_data_recalculate():
    try:
        STATE['loaded'] = False
        STATE['_load_error'] = None
        saved_overrides = dict(STATE['overrides'])
        _run_all()
        # F-05 fix: Wrap remaining STATE mutations in lock
        with STATE_LOCK:
            STATE['overrides'] = saved_overrides
            _apply_all_overrides()
            _recompute_downstream_unlocked()
        return jsonify({'status': 'ok', 'message': 'All engines recalculated',
                        'data': _build_demo_object(), 'initiatives': STATE['initiatives'],
                        'waterfall': STATE['waterfall']})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': 'Internal server error'}), 500


# ══════════════════════════════════════════════════════════════
#  MAIN ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/api/recommendations/<page_context>')
def api_recommendations(page_context):
    """Get contextual recommendations for a specific page."""
    if not STATE['loaded']: return jsonify({'error': 'Not loaded'}), 503
    recs = get_recommendations(page_context, STATE['data'], STATE['diagnostic'],
                                STATE['initiatives'], STATE['waterfall'], maturity=STATE.get('maturity'))
    return jsonify(recs)

@app.route('/api/initiative-linkage/<page_context>')
def api_initiative_linkage(page_context):
    """v12-#23: Get initiative linkage table for a diagnostic page (roadmap-aligned)."""
    if not STATE['loaded']: return jsonify({'error': 'Not loaded'}), 503
    linkage = get_initiative_linkage(page_context, STATE['data'], STATE['diagnostic'],
                                     STATE['initiatives'], STATE['waterfall'])
    return jsonify(linkage)

@app.route('/api/industries')
def api_industries():
    """Get available industry configurations."""
    return jsonify({'industries': get_available_industries()})

@app.route('/api/industry/<key>')
def api_industry_config(key):
    """Get specific industry configuration."""
    config = get_industry_config(key)
    if not config: return jsonify({'error': f'Unknown industry: {key}'}), 404
    return jsonify(config)


# ══════════════════════════════════════════════════════════════
#  V6: MODE TOGGLE & PILLAR SCENARIOS API
# ══════════════════════════════════════════════════════════════

@app.route('/api/config/mode', methods=['GET'])
def api_get_mode():
    """Get current project mode."""
    return jsonify({'mode': get_project_mode()})


@app.route('/api/config/mode', methods=['POST'])
@require_role('admin')
def api_set_mode():
    """Set project mode (admin only)."""
    new_mode = request.json.get('mode', '')
    if new_mode not in ('opportunity', 'delivery'):
        return jsonify({'error': 'Mode must be "opportunity" or "delivery"'}), 400
    user = get_current_user()
    user_id = user['id'] if user else None
    try:
        old_mode = set_project_mode(new_mode, user_id)
        return jsonify({'mode': new_mode, 'previous': old_mode, 'ok': True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/config/mode-log', methods=['GET'])
@require_role('admin')
def api_mode_log():
    """Get mode change audit trail (admin only)."""
    return jsonify({'log': get_mode_change_log()})


@app.route('/api/pillar-scenarios')
def api_pillar_scenarios():
    """Get pillar-isolated waterfall results and ranges."""
    if not STATE['loaded']:
        return jsonify({'error': 'Not loaded'}), 503
    ps = STATE.get('pillarScenarios')
    if not ps:
        return jsonify({'error': 'Pillar scenarios not computed'}), 503
    # Return only the summary and ranges (full waterfall per pillar is too large)
    return jsonify(_sanitize_for_json({
        'ranges': ps['ranges'],
        'pillarSummary': ps['pillarSummary'],
    }))


# ══════════════════════════════════════════════════════════════
#  MAIN PAGE ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if not STATE['loaded'] and not STATE.get('_load_error'):
        try:
            _run_all()
            print("[OK] ContactIQ engines loaded successfully")
        except Exception as e:
            err_msg = f"{type(e).__name__}: {e}"
            STATE['_load_error'] = err_msg
            print(f"\n{'='*60}\n[!] ENGINE LOAD FAILED\n[!] Error: {err_msg}\n{'='*60}\n")
            traceback.print_exc()

    user = get_current_user()
    user_info = {'username': user.get('username','admin'), 'display_name': user.get('display_name','Administrator'),
                 'role': user.get('role','supervisor')} if user else {'username':'admin','display_name':'Administrator','role':'supervisor'}
    data_summary = get_upload_summary()

    # V6: Include project mode and pillar scenarios
    project_mode = get_project_mode()

    if STATE['loaded']:
        # Build pillar summary for frontend (lightweight — not full waterfall per pillar)
        ps = STATE.get('pillarScenarios')
        pillar_data = {
            'ranges': ps['ranges'],
            'pillarSummary': ps['pillarSummary'],
        } if ps else None

        server_data = {
            'demo': _build_demo_object(),
            'initiatives': STATE['initiatives'],
            'waterfall': STATE['waterfall'],
            'pillarScenarios': pillar_data,
            'projectMode': project_mode,
        }
        return render_template('index.html', server_data=json.dumps(server_data, default=str),
                               load_error=None, user_info=json.dumps(user_info),
                               data_summary=json.dumps(data_summary))
    return render_template('index.html', server_data=None, load_error=STATE.get('_load_error', None),
                           user_info=json.dumps(user_info), data_summary=json.dumps(data_summary))


# ══════════════════════════════════════════════════════════════
#  ENGINE API ROUTES (preserved from v7)
# ══════════════════════════════════════════════════════════════

@app.route('/api/data')
def api_data():
    if not STATE['loaded']:
        return jsonify({'error': 'Data not loaded', 'reason': STATE.get('_load_error', 'Unknown')}), 503
    bu = request.args.get('bu')
    demo = _build_demo_object()
    if bu and bu != 'all':
        demo = copy.deepcopy(demo)
        if 'queues' in demo:
            demo['queues'] = [q for q in demo['queues'] if q.get('bu') == bu]
        filtered_q = demo.get('queues', [])
        if filtered_q:
            demo['totalVolumeAnnual'] = sum(q.get('volume', 0) for q in filtered_q)
            demo['totalVolume'] = demo['totalVolumeAnnual']
    
    # V9: Client-role content filtering — strip internal-only fields
    user = get_current_user()
    if user and user.get('role') == 'client':
        demo = copy.deepcopy(demo) if bu is None else demo
        # Strip model internals
        for key in ['auditTrail', 'poolUtilization', 'poolSummary', 'metricSources',
                     'calculationBasis', 'validationIssues', 'subIntentAnalysis']:
            demo.pop(key, None)
        # Strip initiative internal fields
        internal_fields = ['_mechanism', '_grossFTE', '_grossSaving', '_poolConsumed',
                          '_poolCapped', '_capApplied', '_floorApplied', 'matchScore',
                          'relevanceScore', 'relevanceReasons', 'consultantNote']
        for init in demo.get('initiatives', []):
            for f in internal_fields:
                init.pop(f, None)
        # Only show approved initiatives if any have approvalStatus set
        inits = demo.get('initiatives', [])
        has_approval = any(i.get('approvalStatus') for i in inits)
        if has_approval:
            demo['initiatives'] = [i for i in inits if i.get('approvalStatus') == 'approved' or i.get('enabled')]
    
    return jsonify(demo)

@app.route('/api/diagnostic')
def api_diagnostic():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    bu = request.args.get('bu')
    if bu and bu != 'all':
        diag = copy.deepcopy(STATE['diagnostic'])
        # Filter queue-level data by BU
        if 'queueScores' in diag:
            diag['queueScores'] = [q for q in diag['queueScores'] if q.get('bu') == bu]
        if 'subIntentAnalysis' in diag:
            diag['subIntentAnalysis'] = [s for s in diag['subIntentAnalysis']
                                          if any(q.get('bu') == bu for q in STATE['diagnostic'].get('queueScores', [])
                                                 if q.get('intent') == s.get('intent'))]
        # Recompute summary from filtered queues
        qs = diag.get('queueScores', [])
        if qs:
            scores = [q.get('overallScore', 0) for q in qs]
            diag['summary'] = {
                'total': len(qs),
                'avgScore': round(sum(scores) / len(scores), 1) if scores else 0,
                'red': sum(1 for q in qs if q.get('rating') == 'red'),
                'amber': sum(1 for q in qs if q.get('rating') == 'amber'),
                'green': sum(1 for q in qs if q.get('rating') == 'green'),
            }
        return jsonify(diag)
    return jsonify(STATE['diagnostic'])

@app.route('/api/maturity')
def api_maturity():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['maturity'])

@app.route('/api/channel-strategy')
def api_channel_strategy():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['channelStrategy'])

@app.route('/api/initiatives')
def api_initiatives():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify({'initiatives': STATE['initiatives'],
                    'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
                    'totalCount': len(STATE['initiatives'])})

@app.route('/api/waterfall')
def api_waterfall():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    bu = request.args.get('bu')
    if bu and bu != 'all':
        wf = copy.deepcopy(STATE['waterfall'])
        # Scope to BU summary if available
        bu_data = wf.get('buSummary', {}).get(bu, {})
        if bu_data:
            wf['_buScoped'] = bu
            wf['_buData'] = bu_data
        return jsonify(wf)
    return jsonify(STATE['waterfall'])

@app.route('/api/risk')
def api_risk():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['risk'])

@app.route('/api/workforce')
def api_workforce():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    bu = request.args.get('bu')
    if bu and bu != 'all':
        wk = copy.deepcopy(STATE['workforce'])
        # Filter byBU to scoped BU
        if 'byBU' in wk and isinstance(wk['byBU'], dict):
            scoped = {k: v for k, v in wk['byBU'].items() if k == bu}
            wk['byBU'] = scoped
            wk['_buScoped'] = bu
        return jsonify(wk)
    return jsonify(STATE['workforce'])

@app.route('/api/refresh', methods=['POST'])
@require_role('supervisor')
def api_refresh():
    try:
        STATE['overrides'] = {}; STATE['loaded'] = False; STATE['_load_error'] = None
        _run_all()
        return jsonify({'status':'ok', 'data': _build_demo_object(),
                        'initiatives': STATE['initiatives'], 'waterfall': STATE['waterfall']})
    except Exception as e:
        return jsonify({'status':'error','message':'Internal server error'}), 500

@app.route('/api/recalculate', methods=['POST'])
@require_role('supervisor')
def api_recalculate():
    try:
        body = request.get_json(force=True) if request.is_json else {}
        # F-05 fix: Acquire lock BEFORE mutating STATE
        with STATE_LOCK:
            data = STATE['data']
            for key, value in body.get('params', {}).items():
                if key in data['params']:
                    data['params'][key] = value; STATE['overrides'][key] = value
            if 'strategicDriver' in body.get('params', {}):
                data['params']['strategicDriver'] = body['params']['strategicDriver']
            _recompute_all_from_diagnostic_unlocked()
        active_layer = body.get('activeLayer')
        if active_layer and active_layer != 'All Layers':
            scoped_inits = copy.deepcopy(STATE['initiatives'])
            for init in scoped_inits:
                if init.get('layer') != active_layer: init['enabled'] = False
            sw = run_waterfall(data, scoped_inits)
            sr = run_risk(scoped_inits, data)
            swf = run_workforce(data, sw, scoped_inits)
            return jsonify({'status':'ok','scoped':True,'activeLayer':active_layer,
                            'data':_build_demo_object(overrides={'waterfall':sw,'risk':sr,'workforce':swf,'initiatives':scoped_inits}),
                            'initiatives':scoped_inits,'waterfall':sw,'risk':sr,'workforce':swf,
                            'enabledCount':sum(1 for i in scoped_inits if i.get('enabled'))})
        return jsonify({'status':'ok','scoped':False,'data':_build_demo_object(),
                        'initiatives':STATE['initiatives'],'waterfall':STATE['waterfall'],
                        'risk':STATE['risk'],'workforce':STATE['workforce'],
                        'enabledCount':sum(1 for i in STATE['initiatives'] if i.get('enabled')),
                        'delta':STATE.get('lastDelta',{})})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status':'error','message':'Internal server error'}), 500

@app.route('/api/initiative/toggle', methods=['POST'])
@require_role('supervisor')
def api_toggle_initiative():
    body = request.get_json(force=True)
    init_id = body.get('id'); enabled = body.get('enabled')
    if not init_id or enabled is None: return jsonify({'error':'id and enabled required'}), 400
    # F-05 fix: Acquire lock BEFORE mutating STATE to prevent race conditions
    with STATE_LOCK:
        STATE['overrides'][f"init_enabled_{init_id}"] = bool(enabled)
        for init in STATE['initiatives']:
            if init['id'] == init_id: init['enabled'] = bool(enabled); break
        _recompute_downstream_unlocked()
    return jsonify({'status':'ok','enabledCount':sum(1 for i in STATE['initiatives'] if i.get('enabled')),
                    'waterfall':STATE['waterfall'],'risk':STATE['risk'],
                    'workforce':STATE['workforce'],'initiatives':STATE['initiatives']})

@app.route('/api/initiative/update', methods=['POST'])
@require_role('supervisor')
def api_update_initiative():
    body = request.get_json(force=True)
    init_id = body.get('id'); fields = body.get('fields', {})
    if not init_id: return jsonify({'error':'id required'}), 400
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        fk = f"init_fields_{init_id}"
        if fk not in STATE['overrides']: STATE['overrides'][fk] = {}
        STATE['overrides'][fk].update(fields)
        for init in STATE['initiatives']:
            if init['id'] == init_id:
                for k, v in fields.items(): init[k] = v
                break
        _recompute_downstream_unlocked()
    return jsonify({'status':'ok','enabledCount':sum(1 for i in STATE['initiatives'] if i.get('enabled')),
                    'waterfall':STATE['waterfall'],'risk':STATE['risk'],
                    'workforce':STATE['workforce'],'initiatives':STATE['initiatives']})

# V7: Pillar toggle — enables/disables all initiatives in a layer
@app.route('/api/pillar/toggle', methods=['POST'])
@require_role('supervisor')
def api_toggle_pillar():
    body = request.get_json(force=True)
    layer = body.get('layer')
    enabled = body.get('enabled')
    if not layer or enabled is None:
        return jsonify({'error': 'layer and enabled required'}), 400
    valid_layers = ['AI & Automation', 'Operating Model', 'Location Strategy']
    if layer not in valid_layers:
        return jsonify({'error': f'Invalid layer: {layer}'}), 400

    user = get_current_user()
    role = user.get('role', 'supervisor') if user else 'supervisor'
    user_id = user.get('id') if user else None

    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        # Toggle all initiatives in this layer
        toggled = 0
        for init in STATE['initiatives']:
            if init.get('layer') == layer:
                init['enabled'] = bool(enabled)
                STATE['overrides'][f"init_enabled_{init['id']}"] = bool(enabled)
                toggled += 1

        # Store pillar toggle state with role tag for GDS visibility
        STATE['overrides'][f"pillar_toggle_{layer}"] = {
            'enabled': bool(enabled),
            'set_by_role': role,
            'set_by_name': user.get('display_name', '') if user else '',
        }

        # Persist to DB
        # F-20 note: Spec F.3 says pillar toggles are session-level (reset on logout).
        # Deliberate override: we persist to DB for better UX across sessions.
        # This is intentionally stronger than spec — documented deviation.
        save_overrides(STATE['overrides'], user_id)
        _recompute_downstream_unlocked()

    return jsonify({
        'status': 'ok',
        'layer': layer,
        'enabled': bool(enabled),
        'toggled': toggled,
        'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
        'pillarScenarios': {
            'ranges': STATE.get('pillarScenarios', {}).get('ranges') if STATE.get('pillarScenarios') else None,
            'pillarSummary': STATE.get('pillarScenarios', {}).get('pillarSummary') if STATE.get('pillarScenarios') else None,
        },
        'waterfall': STATE['waterfall'],
    })


# V7: Get pillar toggle overrides (for GDS to see what EY US changed)
@app.route('/api/pillar/overrides')
def api_pillar_overrides():
    if not STATE['loaded']:
        return jsonify({'error': 'Not loaded'}), 503
    pillar_overrides = {}
    for key, val in STATE['overrides'].items():
        if key.startswith('pillar_toggle_') and isinstance(val, dict):
            layer = key.replace('pillar_toggle_', '')
            pillar_overrides[layer] = val
    return jsonify({'overrides': pillar_overrides})


# V7: Reset pillar toggles to engine defaults
@app.route('/api/pillar/reset', methods=['POST'])
@require_role('admin')
def api_reset_pillar_toggles():
    # F-05 fix: Acquire lock for all STATE mutations
    with STATE_LOCK:
        # Remove all pillar toggle overrides
        keys_to_remove = [k for k in STATE['overrides'] if k.startswith('pillar_toggle_')]
        for k in keys_to_remove:
            del STATE['overrides'][k]
        # Re-score initiatives from scratch (resets enables to engine defaults)
        from engines.waterfall import score_initiatives
        STATE['initiatives'] = score_initiatives(STATE['data'], STATE['diagnostic'], STATE['readiness'])
        _apply_all_overrides()  # Re-apply non-pillar overrides
        save_overrides(STATE['overrides'])
        _recompute_downstream_unlocked()
    return jsonify({
        'status': 'ok',
        'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
    })


@app.route('/api/override', methods=['POST'])
@require_role('supervisor')
def api_override():
    body = request.get_json(force=True)
    key = body.get('key'); value = body.get('value')
    if not key: return jsonify({'error':'key required'}), 400
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        STATE['overrides'][key] = value
        if key in STATE['data']['params']: STATE['data']['params'][key] = value
        _recompute_all_from_diagnostic_unlocked()
    return jsonify({'status':'ok','data':_build_demo_object(),
                    'initiatives':STATE['initiatives'],'waterfall':STATE['waterfall']})

@app.route('/api/subintent/override', methods=['POST'])
@require_role('supervisor')
def api_subintent_override():
    """CR-09: Persist sub-intent level overrides and trigger full recalculation."""
    body = request.get_json(force=True)
    intent = body.get('intent'); subintent = body.get('subintent')
    if not intent or not subintent:
        return jsonify({'error': 'intent and subintent required'}), 400
    okey = f"subintent_{intent}_{subintent}"
    fields = {}
    for f in ('volShare', 'complexity', 'lever', 'deflectable', 'fteOverride'):
        if f in body:
            fields[f] = body[f]
    if not fields:
        return jsonify({'error': 'no override fields provided'}), 400
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        if okey not in STATE['overrides']:
            STATE['overrides'][okey] = {}
        STATE['overrides'][okey].update(fields)
        _recompute_all_from_diagnostic_unlocked()
    return jsonify({
        'status': 'ok',
        'data': _build_demo_object(),
        'initiatives': STATE['initiatives'],
        'waterfall': STATE['waterfall']
    })

@app.route('/api/maturity/override', methods=['POST'])
@require_role('supervisor')
def api_maturity_override():
    body = request.get_json(force=True)
    dimension = body.get('dimension'); score = body.get('score')
    if not dimension or score is None: return jsonify({'error':'dimension and score required'}), 400
    score = max(1.0, min(5.0, float(score)))
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        STATE['overrides'][f"maturity_{dimension}"] = score
        mat = STATE['maturity']; dims = mat.get('dimensions', {})
        if dimension in dims:
            dims[dimension]['score'] = score; dims[dimension]['level'] = min(5, max(1, round(score)))
        if dims:
            from engines.maturity import DIMENSIONS, MATURITY_LEVELS
            overall = sum(dims[k]['score'] * dims[k].get('weight', 0.20) for k in dims if isinstance(dims[k], dict))
            mat['overall'] = round(overall, 2)
            mat['overallLevel'] = min(5, max(1, round(overall)))
            mat['levelInfo'] = MATURITY_LEVELS.get(mat['overallLevel'], {})
    return jsonify({'status':'ok','maturity': mat})

@app.route('/api/benchmarks/override', methods=['POST'])
@require_role('supervisor')
def api_benchmark_override():
    """Save benchmark overrides persistently across sessions."""
    body = request.get_json(force=True)
    overrides = body.get('benchmarks', {})
    if not overrides: return jsonify({'error': 'benchmarks object required'}), 400
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        for metric, value in overrides.items():
            STATE['overrides'][f"benchmark_{metric}"] = value
            if metric in STATE['data'].get('benchmarks', {}):
                if isinstance(STATE['data']['benchmarks'][metric], dict):
                    STATE['data']['benchmarks'][metric]['global'] = value
                else:
                    STATE['data']['benchmarks'][metric] = value
        _recompute_all_from_diagnostic_unlocked()
    return jsonify({'status': 'ok', 'message': f'{len(overrides)} benchmark(s) saved',
                    'data': _build_demo_object(), 'waterfall': STATE['waterfall']})

@app.route('/api/benchmarks/overrides', methods=['GET'])
def api_benchmark_overrides_get():
    """Retrieve saved benchmark overrides."""
    bm_overrides = {k.replace('benchmark_', ''): v for k, v in STATE['overrides'].items() if k.startswith('benchmark_')}
    return jsonify({'benchmarks': bm_overrides})


@app.route('/api/operating-model/save', methods=['POST'])
@require_role('supervisor')
def api_operating_model_save():
    """A-05 fix: Persist target operating model to STATE overrides."""
    body = request.get_json(force=True)
    om = body.get('operatingModel', {})
    if not om:
        return jsonify({'error': 'operatingModel object required'}), 400
    STATE['overrides']['operating_model'] = om
    return jsonify({'status': 'ok', 'message': 'Operating model saved'})


@app.route('/api/operating-model/load', methods=['GET'])
def api_operating_model_load():
    """A-05 fix: Load persisted target operating model."""
    om = STATE['overrides'].get('operating_model', {})
    return jsonify({'operatingModel': om})


# ── V9: Per-page consultant narrative ──
@app.route('/api/narrative/<int:page>', methods=['GET'])
def api_get_narrative(page):
    key = f'page_narrative_{page}'
    text = STATE['overrides'].get(key, '')
    return jsonify({'page': page, 'text': text})

@app.route('/api/narrative/<int:page>', methods=['POST'])
@require_role('supervisor')
def api_set_narrative(page):
    body = request.get_json(force=True)
    text = body.get('text', '').strip()
    key = f'page_narrative_{page}'
    with STATE_LOCK:
        STATE['overrides'][key] = text
        save_overrides(STATE['overrides'], user_id=get_current_user().get('id') if get_current_user() else None, reason=f'Narrative update for page {page}')
    return jsonify({'status': 'ok', 'page': page})

@app.route('/api/narratives', methods=['GET'])
def api_get_all_narratives():
    narratives = {k: v for k, v in STATE['overrides'].items() if k.startswith('page_narrative_')}
    return jsonify({'narratives': narratives})

@app.route('/api/investment')
def api_investment():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    wf = STATE['waterfall']
    return jsonify({'items':wf.get('investmentItems',[]),'summary':wf.get('investmentSummary',{}),'yearly':wf.get('investmentYearly',[])})

@app.route('/api/initiatives/batch', methods=['POST'])
@require_role('supervisor')
def api_batch_initiatives():
    body = request.get_json(force=True)
    # F-05 fix: Acquire lock BEFORE mutating STATE
    with STATE_LOCK:
        for upd in body.get('updates', []):
            iid = upd.get('id')
            for init in STATE['initiatives']:
                if init['id'] == iid:
                    if 'enabled' in upd:
                        init['enabled'] = bool(upd['enabled']); STATE['overrides'][f"init_enabled_{iid}"] = bool(upd['enabled'])
                    for rk in ('rampYear1','rampYear2','rampYear3'):
                        if rk in upd: init[rk] = float(upd[rk]); STATE['overrides'][f"init_{rk}_{iid}"] = float(upd[rk])
                    if 'priority' in upd: init['priority'] = upd['priority']
                    break
        _recompute_downstream_unlocked()
    return jsonify({'status':'ok','initiatives':STATE['initiatives'],'waterfall':STATE['waterfall']})

@app.route('/api/export')
def api_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()

        # ── Style definitions ──
        ey_dark = '2E2E38'
        ey_yellow = 'FFE600'
        ey_green = '168736'
        ey_red = 'C5003E'
        ey_amber = 'D97706'

        hf = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
        hfill = PatternFill(start_color=ey_dark, end_color=ey_dark, fill_type='solid')
        title_font = Font(bold=True, color=ey_dark, size=14, name='Calibri')
        subtitle_font = Font(bold=False, color='747480', size=10, name='Calibri')
        data_font = Font(color=ey_dark, size=10, name='Calibri')
        bold_font = Font(bold=True, color=ey_dark, size=10, name='Calibri')
        tb = Border(left=Side(style='thin', color='E0E0E6'), right=Side(style='thin', color='E0E0E6'),
                    top=Side(style='thin', color='E0E0E6'), bottom=Side(style='thin', color='E0E0E6'))
        alt_fill = PatternFill(start_color='F6F6FA', end_color='F6F6FA', fill_type='solid')
        green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
        amber_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')
        rag_fills = {'green': green_fill, 'amber': amber_fill, 'red': red_fill}

        def ws_write(ws, headers, rows, start_row=1, col_widths=None):
            # Headers
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=c, value=h)
                cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = tb
            ws.row_dimensions[start_row].height = 26
            # Data rows
            for r, row in enumerate(rows, start_row + 1):
                is_alt = (r - start_row) % 2 == 0
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.font = data_font; cell.border = tb
                    cell.alignment = Alignment(vertical='center')
                    if is_alt: cell.fill = alt_fill
            # Auto-width
            if col_widths:
                for i, w in enumerate(col_widths):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w
            else:
                for col_cells in ws.columns:
                    ml = max(len(str(cell.value or '')) for cell in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(ml + 3, 45)
            # Freeze panes
            ws.freeze_panes = ws.cell(row=start_row+1, column=1)

        def add_title(ws, title, subtitle='', row=1):
            ws.cell(row=row, column=1, value=title).font = title_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            if subtitle:
                ws.cell(row=row+1, column=1, value=subtitle).font = subtitle_font
                ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=6)
                return row + 3
            return row + 2

        data = STATE['data']; diag = STATE['diagnostic']; wf = STATE['waterfall']
        params = data['params']

        # ══════ Sheet 1: Executive Summary ══════
        ws = wb.active; ws.title = 'Executive Summary'
        ws.sheet_properties.tabColor = ey_yellow
        sr = add_title(ws, f"ContactIQ — {params.get('clientName','Client')}", f"{params.get('industry','')} | Generated Report")
        summary_data = [
            ['Total Annual Volume', data.get('totalVolumeAnnual', data['totalVolume'])],
            ['Total FTE', data['totalFTE']],
            ['Total Annual Cost', data['totalCost']],
            ['Average Cost Per Contact', data.get('avgCPC', 0)],
            ['', ''],
            ['Net Present Value', wf.get('totalNPV', 0)],
            ['Total 3-Year Saving', wf.get('totalSaving', 0)],
            ['Total Investment', wf.get('totalInvestment', 0)],
            ['ROI', wf.get('roi', 0)],
            ['IRR', wf.get('irr', 0)],
            ['Payback (months)', wf.get('payback', 0)],
            ['Enabled Initiatives', sum(1 for i in STATE['initiatives'] if i.get('enabled'))],
        ]
        for r, (label, val) in enumerate(summary_data, sr):
            ws.cell(row=r, column=1, value=label).font = bold_font
            ws.cell(row=r, column=1).border = tb
            cell = ws.cell(row=r, column=2, value=val)
            cell.border = tb
            if isinstance(val, (int, float)) and val > 1000:
                cell.number_format = '#,##0'
                if label.startswith(('Total Annual Cost', 'Net Present', 'Total 3-Year', 'Total Investment', 'Average Cost')):
                    cell.number_format = '$#,##0'
            if label in ('ROI', 'IRR'):
                cell.number_format = '0.0"%"'
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 22

        # ══════ Sheet 2: KPI Projections ══════
        kpis = wf.get('kpiProjections', {})
        if kpis:
            wskpi = wb.create_sheet('KPI Projections')
            wskpi.sheet_properties.tabColor = '1A73E8'
            ksr = add_title(wskpi, 'Operational KPI Improvement', 'Current → Projected → Benchmark with RAG')
            kpi_headers = ['KPI', 'Current', 'Projected', 'Benchmark', 'Delta', 'Delta %', 'RAG', 'Top Contributors']
            kpi_rows = []
            kpi_rags = []
            for k in ['AHT', 'FCR', 'CSAT', 'CPC', 'Escalation', 'Repeat', 'CES']:
                kp = kpis.get(k)
                if not kp: continue
                contribs = ', '.join(c.get('name','')[:25] for c in (kp.get('contributors') or [])[:3])
                kpi_rows.append([kp['label'], kp['current'], kp['projected'], kp['benchmark'],
                                 kp['delta'], kp['deltaPct'], kp.get('rag','').upper(), contribs or 'Derived'])
                kpi_rags.append(kp.get('rag', ''))
            ws_write(wskpi, kpi_headers, kpi_rows, start_row=ksr, col_widths=[24, 14, 14, 14, 12, 12, 10, 36])
            # Apply RAG coloring
            for i, rag in enumerate(kpi_rags):
                if rag in rag_fills:
                    for c in range(1, 9):
                        wskpi.cell(row=ksr+1+i, column=c).fill = rag_fills[rag]

        # ══════ Sheet 3: Initiatives ══════
        ws5 = wb.create_sheet('Initiatives')
        ws5.sheet_properties.tabColor = ey_green
        isr = add_title(ws5, 'Initiative Portfolio', f'{sum(1 for i in STATE["initiatives"] if i.get("enabled"))} enabled of {len(STATE["initiatives"])}')
        init_rows = []
        for i in STATE['initiatives']:
            init_rows.append([
                i['id'], i['name'], i['layer'], i['lever'],
                ', '.join(sorted(i.get('levers', {i.get('lever', ''): True}).keys())),
                'Yes' if i.get('enabled') else 'No',
                round(i.get('matchScore', 0), 1),
                round(i.get('_annualSaving', 0)),
                round(i.get('_fteImpact', 0), 1),
                i.get('implRisk', ''),
                i.get('cxRisk', ''),
                i.get('opsRisk', ''),
                i.get('overallRisk', ''),
                i.get('riskRating', ''),
            ])
        ws_write(ws5, ['ID', 'Name', 'Layer', 'Lever', 'Levers (All)', 'Enabled', 'Score',
                        'Annual Saving ($)', 'FTE Impact', 'Impl Risk', 'CX Risk', 'Ops Risk', 'Overall Risk', 'Rating'],
                 init_rows, start_row=isr, col_widths=[6, 34, 18, 18, 24, 10, 10, 18, 12, 10, 10, 10, 12, 10])
        # Format savings column as currency + color enabled/disabled
        for r in range(isr+1, isr+1+len(init_rows)):
            ws5.cell(row=r, column=8).number_format = '$#,##0'
            # Color enabled/disabled
            en_cell = ws5.cell(row=r, column=6)
            if en_cell.value == 'Yes':
                en_cell.fill = green_fill
            else:
                en_cell.fill = red_fill

        # ══════ Sheet 4: Waterfall ══════
        ws6 = wb.create_sheet('Waterfall')
        ws6.sheet_properties.tabColor = '4CAF50'
        wsr = add_title(ws6, 'Yearly Financial Projections')
        wf_rows = [[y['year'], round(y['annualSaving']), round(y['cumSaving']), round(y['npv']),
                     round(y.get('futureCost', 0)), y.get('finalFTE', 0)]
                    for y in wf.get('yearly', [])]
        ws_write(ws6, ['Year', 'Annual Saving ($)', 'Cumulative ($)', 'NPV ($)', 'Future Cost ($)', 'Final FTE'],
                 wf_rows, start_row=wsr, col_widths=[10, 20, 20, 20, 20, 14])
        for r in range(wsr+1, wsr+1+len(wf_rows)):
            for c in [2,3,4,5]:
                ws6.cell(row=r, column=c).number_format = '$#,##0'

        # ══════ Sheet 5: Channel Mix ══════
        ch_mix = _build_channel_mix(data['queues'])
        if ch_mix:
            wsch = wb.create_sheet('Channel Mix')
            wsch.sheet_properties.tabColor = '0097A9'
            csr = add_title(wsch, 'Channel Volume & Performance')
            ch_rows = [[c['channel'], c['volume'], c['pct'], c.get('avgCSAT', 0), c.get('avgAHT_min', 0), c.get('avgCPC', 0)]
                       for c in ch_mix]
            ws_write(wsch, ['Channel', 'Volume', 'Share %', 'Avg CSAT', 'Avg AHT (min)', 'Avg CPC ($)'],
                     ch_rows, start_row=csr, col_widths=[20, 14, 12, 12, 14, 14])
            for r in range(csr+1, csr+1+len(ch_rows)):
                wsch.cell(row=r, column=2).number_format = '#,##0'
                wsch.cell(row=r, column=3).number_format = '0.0"%"'
                wsch.cell(row=r, column=6).number_format = '$#,##0.00'

        # ══════ Sheet 6: BU Impact ══════
        bu_summary = wf.get('buSummary', {})
        if bu_summary:
            wsbu = wb.create_sheet('BU Impact')
            wsbu.sheet_properties.tabColor = ey_amber
            bsr = add_title(wsbu, 'Business Unit Impact')
            bu_rows = []
            for bu, bd in bu_summary.items():
                for yr_idx, yr_data in enumerate(bd.get('yearly', [])):
                    bu_rows.append([bu, yr_idx+1, round(yr_data.get('annualSaving', 0))])
            ws_write(wsbu, ['Business Unit', 'Year', 'Annual Saving ($)'], bu_rows, start_row=bsr, col_widths=[24, 10, 20])
            for r in range(bsr+1, bsr+1+len(bu_rows)):
                wsbu.cell(row=r, column=3).number_format = '$#,##0'

        # ══════ Sheet 7: Risk Register ══════
        risk_data = STATE.get('risk', {})
        risk_inits = risk_data.get('initiatives', [])
        if risk_inits:
            wsrisk = wb.create_sheet('Risk Register')
            wsrisk.sheet_properties.tabColor = ey_red
            rsr = add_title(wsrisk, 'Risk Assessment')
            risk_rows = [[r['id'], r['name'], r['layer'], r['implRisk'], r['cxRisk'], r['opsRisk'],
                          r['overallRisk'], r['rating'], round(r['annualSaving']),
                          '; '.join(r.get('mitigations', []))]
                         for r in risk_inits]
            ws_write(wsrisk, ['ID', 'Name', 'Layer', 'Impl', 'CX', 'Ops', 'Overall', 'Rating', 'Saving ($)', 'Mitigations'],
                     risk_rows, start_row=rsr, col_widths=[6, 30, 18, 8, 8, 8, 10, 10, 16, 40])
            for r_idx in range(rsr+1, rsr+1+len(risk_rows)):
                wsrisk.cell(row=r_idx, column=9).number_format = '$#,##0'
                rating_cell = wsrisk.cell(row=r_idx, column=8)
                rv = str(rating_cell.value).lower()
                if rv == 'low': rating_cell.fill = green_fill
                elif rv == 'medium': rating_cell.fill = amber_fill
                elif rv == 'high': rating_cell.fill = red_fill

        # ══════ Sheet 8: Workforce Transition ══════
        wkf = STATE.get('workforce', {})
        transitions = wkf.get('transitions', [])
        if transitions:
            wswf = wb.create_sheet('Workforce Transition')
            wsr2 = add_title(wswf, 'Workforce Transition Plan')
            wf_t_rows = [[t.get('role',''), t.get('location',''), t.get('sourcing',''),
                          t.get('year',0), round(t.get('reduction',0),1), round(t.get('attrited',0),1),
                          round(t.get('redeployed',0),1), round(t.get('separated',0),1),
                          round(t.get('totalTransitionCost',0))]
                         for t in transitions]
            ws_write(wswf, ['Role', 'Location', 'Sourcing', 'Year', 'Reduction', 'Attrited', 'Redeployed', 'Separated', 'Transition Cost ($)'],
                     wf_t_rows, start_row=wsr2, col_widths=[18, 14, 14, 8, 12, 12, 12, 12, 18])
            for r in range(wsr2+1, wsr2+1+len(wf_t_rows)):
                wswf.cell(row=r, column=9).number_format = '$#,##0'

        # ══════ Sheet 9: Location Cost Matrix ══════
        loc_matrix = data.get('locationCostMatrix', {})
        if loc_matrix:
            wsloc = wb.create_sheet('Location Cost Matrix')
            wsloc.sheet_properties.tabColor = '7B61FF'
            lsr = add_title(wsloc, 'Location & Sourcing Cost Matrix')
            loc_rows = []
            for loc, srcs in loc_matrix.items():
                for src, costs in srcs.items():
                    loc_rows.append([loc, src, costs.get('costPerFTE', 0),
                                     costs.get('hiringCost', 0),
                                     costs.get('attritionRate', 0)])
            ws_write(wsloc, ['Location', 'Sourcing', 'Cost/FTE ($)', 'Hiring Cost ($)', 'Attrition Rate'],
                     loc_rows, start_row=lsr, col_widths=[18, 18, 16, 16, 16])
            for r in range(lsr+1, lsr+1+len(loc_rows)):
                wsloc.cell(row=r, column=3).number_format = '$#,##0'
                wsloc.cell(row=r, column=4).number_format = '$#,##0'
                wsloc.cell(row=r, column=5).number_format = '0.0%'

        fd, export_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(export_path)
        return send_file(export_path, as_attachment=True, download_name=f"ContactIQ_{params.get('clientName', 'Client')}_Export.xlsx")
    except ImportError as e:
        return jsonify({'error': f'Export dependency missing: {e}. Install with: pip install openpyxl'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error':'Internal server error'}), 500


# ══════════════════════════════════════════════════════════════
#  LAYER FILTER — Server-side recompute (Option A)
# ══════════════════════════════════════════════════════════════

@app.route('/api/waterfall/layer', methods=['POST'])
def api_waterfall_by_layer():
    """Recompute waterfall with only initiatives from specified layer(s)."""
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    try:
        body = request.get_json(force=True)
        active_layer = body.get('layer', 'All Layers')
        
        if active_layer == 'All Layers':
            return jsonify({
                'status': 'ok', 'scoped': False, 'layer': 'All Layers',
                'waterfall': STATE['waterfall'],
                'initiatives': STATE['initiatives'],
                'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
            })
        
        scoped_inits = copy.deepcopy(STATE['initiatives'])
        for init in scoped_inits:
            if init.get('layer') != active_layer:
                init['enabled'] = False
        
        data = STATE['data']
        sw = run_waterfall(data, scoped_inits)
        
        return jsonify({
            'status': 'ok', 'scoped': True, 'layer': active_layer,
            'waterfall': sw,
            'initiatives': scoped_inits,
            'enabledCount': sum(1 for i in scoped_inits if i.get('enabled')),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500


# ══════════════════════════════════════════════════════════════
#  PDF EXPORT
# ══════════════════════════════════════════════════════════════

def _pdf_safe(text):
    """Sanitize text for fpdf2 Helvetica (Latin-1 only) — replaces Unicode chars that cause crashes."""
    return str(text).replace('\u2014', '-').replace('\u2013', '-').replace('\u2018', "'").replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"').replace('\u2026', '...').replace('\u2192', '->').replace('\u2190', '<-').replace('\u2022', '*').replace('\u2713', 'Y').replace('\u2717', 'X').replace('\u2605', '*').replace('\u2714', 'Y').replace('\u2716', 'X').replace('\u25CF', '*').replace('\u2019', "'").replace('\u00a0', ' ')

@app.route('/api/export/pdf')
def api_export_pdf():
    """V7.1: Role-aware PDF report — content matches page access per role/mode."""
    try:
        from fpdf import FPDF

        data = STATE['data']; wf = STATE['waterfall']; diag = STATE['diagnostic']
        params = data['params']

        # Get current user role and mode
        # F-21 note: Per spec F.2.8, mode is read ONCE at the START of PDF generation
        # and used for the entire document. No mid-generation mode check.
        # If mode later moves to DB-backed with concurrent writes, snapshot this value.
        user = get_current_user()
        role = user.get('role', 'admin') if user else 'admin'
        mode = get_project_mode()

        # Section visibility per Appendix E
        def section_allowed(section):
            matrix = {
                'financial_overview': {
                    'admin': True,
                    'supervisor': True,
                    'analyst': mode == 'delivery',
                },
                'initiatives': {
                    'admin': True,
                    'supervisor': mode == 'delivery',
                    'analyst': False,
                },
                'pools': {
                    'admin': True,
                    'supervisor': False,
                    'analyst': False,
                },
                'risk': {
                    'admin': True,
                    'supervisor': mode == 'delivery',
                    'analyst': mode == 'delivery',
                },
                'maturity': {
                    'admin': True,
                    'supervisor': True,
                    'analyst': False,
                },
                'channel_detail': {
                    'admin': True,
                    'supervisor': True,
                    'analyst': False,
                },
                'diagnostic_detail': {
                    'admin': True,
                    'supervisor': True,
                    'analyst': False,
                },
                'channel_migration': {
                    'admin': True,
                    'supervisor': False,
                    'analyst': False,
                },
                'operating_model': {
                    'admin': True,
                    'supervisor': True,
                    'analyst': False,
                },
            }
            return matrix.get(section, {}).get(role, False)

        # CR-12: Wrap FPDF to auto-sanitize all text for Latin-1 Helvetica
        class SafeFPDF(FPDF):
            def cell(self, w=0, h=0, text='', *args, **kwargs):
                return super().cell(w, h, _pdf_safe(text), *args, **kwargs)
            def multi_cell(self, w, h=0, text='', *args, **kwargs):
                return super().multi_cell(w, h, _pdf_safe(text), *args, **kwargs)

        pdf = SafeFPDF()
        pdf.set_auto_page_break(auto=True, margin=20)

        # ── Cover Page ──
        pdf.add_page()
        pdf.set_fill_color(46, 46, 56)
        pdf.rect(0, 0, 210, 297, 'F')
        pdf.set_text_color(255, 230, 0)
        pdf.set_font('Helvetica', 'B', 32)
        pdf.set_y(80)
        pdf.cell(0, 15, 'Contact Centre', ln=True, align='C')
        pdf.cell(0, 15, 'Opportunity', ln=True, align='C')
        pdf.cell(0, 15, 'Assessment', ln=True, align='C')
        pdf.set_font('Helvetica', '', 14)
        pdf.set_text_color(255, 255, 255)
        pdf.ln(20)
        pdf.cell(0, 10, params.get('clientName', 'Client'), ln=True, align='C')
        pdf.cell(0, 10, params.get('industry', 'Industry'), ln=True, align='C')
        pdf.set_font('Helvetica', '', 10)
        pdf.ln(10)
        pdf.cell(0, 8, 'Generated by ContactIQ — EY Contact Centre Opportunity Assessment', ln=True, align='C')
        # Role/mode watermark
        role_label = {'admin':'EY GDS','supervisor':'EY US','analyst':'Client'}.get(role, role)
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(180, 180, 180)
        pdf.ln(30)
        pdf.cell(0, 8, f'Report prepared for: {role_label} | Mode: {mode.capitalize()}', ln=True, align='C')

        # ── Executive Summary ──
        pdf.add_page()
        pdf.set_text_color(46, 46, 56)
        pdf.set_font('Helvetica', 'B', 18)
        pdf.cell(0, 12, 'Executive Summary', ln=True)
        pdf.set_draw_color(255, 230, 0)
        pdf.set_line_width(1)
        pdf.line(10, pdf.get_y(), 80, pdf.get_y())
        pdf.ln(8)

        if section_allowed('financial_overview'):
            metrics = [
                ('Total Savings', f"${wf.get('totalSaving', 0):,.0f}"),
                ('Net Present Value', f"${wf.get('totalNPV', 0):,.0f}"),
                ('Internal Rate of Return', f"{wf.get('irr', 0):.1f}%"),
                ('Payback Period', f"{wf.get('payback', 0):.1f} years"),
                ('Total Investment', f"${wf.get('totalInvestment', 0):,.0f}"),
            ]
            pdf.set_font('Helvetica', '', 11)
            for label, value in metrics:
                pdf.set_font('Helvetica', 'B', 11)
                pdf.cell(90, 8, label, border='B')
                pdf.set_font('Helvetica', '', 11)
                pdf.cell(0, 8, value, border='B', ln=True)
            pdf.ln(6)

        # Pillar summary (all roles see this)
        ps = STATE.get('pillarScenarios')
        if ps:
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Opportunity by Pillar', ln=True)
            pdf.set_font('Helvetica', '', 10)
            for key, label in [('ai','AI & Automation'),('om','Operating Model'),('location','Location Strategy')]:
                p = ps['pillarSummary'].get(key, {})
                ann = p.get('annualSaving', 0)
                fte = p.get('totalReduction', 0)
                fte_str = f", ~{fte:.0f} FTE reduction" if fte > 0 else " (cost savings only)"
                pdf.cell(0, 7, f"  {label}: ${ann:,.0f}/yr{fte_str}", ln=True)
            ranges = ps.get('ranges', {})
            fr = ranges.get('fteRange', {})
            sr = ranges.get('savingsRange', {})
            pdf.ln(3)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 7, f"FTE Range: {fr.get('low',0):.0f} - {fr.get('high',0):.0f}  |  Savings Range: ${sr.get('low',0):,.0f} - ${sr.get('high',0):,.0f}", ln=True)
            pdf.ln(4)

        # Yearly projections (financial_overview gated)
        if section_allowed('financial_overview'):
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Yearly Projections', ln=True)
            pdf.set_font('Helvetica', 'B', 9)
            headers = ['Year', 'Annual Saving', 'Cumulative', 'NPV']
            col_w = [30, 45, 45, 45]
            pdf.set_fill_color(46, 46, 56)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 8, h, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_text_color(46, 46, 56)
            pdf.set_font('Helvetica', '', 9)
            for y in wf.get('yearly', []):
                pdf.cell(col_w[0], 7, str(y.get('year', '')), border=1, align='C')
                pdf.cell(col_w[1], 7, f"${y.get('annualSaving', 0):,.0f}", border=1, align='C')
                pdf.cell(col_w[2], 7, f"${y.get('cumSaving', 0):,.0f}", border=1, align='C')
                pdf.cell(col_w[3], 7, f"${y.get('npv', 0):,.0f}", border=1, align='C')
                pdf.ln()

        # Layer Breakdown
        if section_allowed('financial_overview'):
            pdf.ln(8)
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Impact by Layer', ln=True)
            pdf.set_font('Helvetica', 'B', 9)
            lh = ['Layer', 'Annual Saving']
            lw = [100, 80]
            pdf.set_fill_color(46, 46, 56)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(lh):
                pdf.cell(lw[i], 8, h, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_text_color(46, 46, 56)
            pdf.set_font('Helvetica', '', 9)
            for layer, fte in wf.get('layerFTE', {}).items():
                saving = wf.get('layerSaving', {}).get(layer, 0)
                pdf.cell(lw[0], 7, layer, border=1)
                pdf.cell(lw[1], 7, f"${saving:,.0f}", border=1, align='C')
                pdf.ln()

        # ── Top Initiatives ──
        if section_allowed('initiatives'):
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Top Initiatives by Impact', ln=True)
            enabled = [i for i in STATE['initiatives'] if i.get('enabled') and i.get('_annualSaving', 0) > 0]
            enabled.sort(key=lambda x: x.get('_annualSaving', 0), reverse=True)
            pdf.set_font('Helvetica', 'B', 8)
            ih = ['#', 'Initiative', 'Layer', 'Lever', 'Annual Saving']
            iw = [8, 65, 35, 40, 40]
            pdf.set_fill_color(46, 46, 56)
            pdf.set_text_color(255, 255, 255)
            for i, h in enumerate(ih):
                pdf.cell(iw[i], 7, h, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_text_color(46, 46, 56)
            pdf.set_font('Helvetica', '', 8)
            for idx, init in enumerate(enabled[:20], 1):
                pdf.cell(iw[0], 6, str(idx), border=1, align='C')
                name = init['name'][:32] + '..' if len(init['name']) > 34 else init['name']
                pdf.cell(iw[1], 6, name, border=1)
                pdf.cell(iw[2], 6, init.get('layer', '')[:18], border=1, align='C')
                pdf.cell(iw[3], 6, init.get('lever', '').replace('_', ' ')[:20], border=1, align='C')
                pdf.cell(iw[4], 6, f"${init.get('_annualSaving', 0):,.0f}", border=1, align='C')
                pdf.ln()

        # ── Pool Utilization ──
        if section_allowed('pools'):
            pdf.ln(8)
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Opportunity Pool Utilization', ln=True)
            pools = wf.get('poolUtilization', wf.get('poolSummary', {}))
            if pools:
                pdf.set_font('Helvetica', 'B', 9)
                ph = ['Pool', 'Ceiling', 'Consumed', 'Remaining', 'Utilization']
                pw = [45, 35, 35, 35, 35]
                pdf.set_fill_color(46, 46, 56)
                pdf.set_text_color(255, 255, 255)
                for i, h in enumerate(ph):
                    pdf.cell(pw[i], 8, h, border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_text_color(46, 46, 56)
                pdf.set_font('Helvetica', '', 9)
                pool_items = pools.items() if isinstance(pools, dict) else enumerate(pools)
                for key, pool in pool_items:
                    if isinstance(pool, dict):
                        ceil_val = pool.get('ceiling_fte', pool.get('ceiling', 0))
                        cons_val = pool.get('consumed_fte', pool.get('consumed', 0))
                        rem_val = pool.get('remaining_fte', pool.get('remaining', ceil_val - cons_val))
                        util_pct = pool.get('utilization_pct', round(cons_val / max(ceil_val, 1) * 100, 1))
                        pdf.cell(pw[0], 7, str(key).replace('_', ' ').title(), border=1)
                        pdf.cell(pw[1], 7, f"{ceil_val:,.1f}", border=1, align='C')
                        pdf.cell(pw[2], 7, f"{cons_val:,.1f}", border=1, align='C')
                        pdf.cell(pw[3], 7, f"{rem_val:,.1f}", border=1, align='C')
                        pdf.cell(pw[4], 7, f"{util_pct:.1f}%", border=1, align='C')
                        pdf.ln()

        # ── Risk Assessment ──
        if section_allowed('risk'):
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 18)
            pdf.set_text_color(46, 46, 56)
            pdf.cell(0, 12, 'Risk Assessment', ln=True)
            pdf.set_draw_color(255, 230, 0)
            pdf.set_line_width(1)
            pdf.line(10, pdf.get_y(), 80, pdf.get_y())
            pdf.ln(8)
            risk_data = STATE.get('risk', {})
            risk_summary = risk_data.get('summary', {})
            overall_risk = risk_summary.get('avgRisk', risk_summary.get('overallScore', 0))
            if overall_risk == 0 and risk_data.get('initiatives'):
                risk_inits_list = risk_data.get('initiatives', [])
                if risk_inits_list:
                    overall_risk = sum(r.get('overallRisk', 0) for r in risk_inits_list) / len(risk_inits_list)
            risk_level = 'Low' if overall_risk < 2 else ('Medium' if overall_risk < 3.5 else 'High')
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(0, 8, f'Overall Risk Score: {overall_risk:.1f}/5 ({risk_level})', ln=True)
            pdf.ln(4)
            risk_dims = risk_data.get('dimensions', {})
            if risk_dims:
                pdf.set_font('Helvetica', 'B', 9)
                rh = ['Dimension', 'Score', 'Level', 'Mitigation']
                rw = [45, 25, 25, 95]
                pdf.set_fill_color(46, 46, 56)
                pdf.set_text_color(255, 255, 255)
                for i, h in enumerate(rh):
                    pdf.cell(rw[i], 8, h, border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_text_color(46, 46, 56)
                pdf.set_font('Helvetica', '', 8)
                for dim_name, dim_data in risk_dims.items():
                    if isinstance(dim_data, dict):
                        s = dim_data.get('score', 0)
                        lvl = 'Low' if s < 2 else ('Medium' if s < 3.5 else 'High')
                        mit = dim_data.get('mitigation', dim_data.get('recommendation', ''))[:60]
                        pdf.cell(rw[0], 7, str(dim_name).replace('_', ' ').title(), border=1)
                        pdf.cell(rw[1], 7, f'{s:.1f}', border=1, align='C')
                        pdf.cell(rw[2], 7, lvl, border=1, align='C')
                        pdf.cell(rw[3], 7, mit, border=1)
                        pdf.ln()

        # ── Maturity Assessment ──
        if section_allowed('maturity'):
            pdf.ln(10)
            pdf.set_font('Helvetica', 'B', 18)
            pdf.cell(0, 12, 'Maturity Assessment', ln=True)
            pdf.set_draw_color(255, 230, 0)
            pdf.line(10, pdf.get_y(), 80, pdf.get_y())
            pdf.ln(8)
            mat = STATE.get('maturity', {})
            mat_overall = mat.get('overall', 0)
            mat_level = mat.get('overallLevel', 0)
            level_info = mat.get('levelInfo', {})
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(0, 8, f'Overall Maturity: {mat_overall:.1f}/5 (Level {mat_level}: {level_info.get("name", "")})', ln=True)
            pdf.ln(4)
            mat_dims = mat.get('dimensions', {})
            if mat_dims:
                pdf.set_font('Helvetica', 'B', 9)
                mh = ['Dimension', 'Score', 'Level', 'Weight']
                mw = [55, 25, 25, 25]
                pdf.set_fill_color(46, 46, 56)
                pdf.set_text_color(255, 255, 255)
                for i, h in enumerate(mh):
                    pdf.cell(mw[i], 8, h, border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_text_color(46, 46, 56)
                pdf.set_font('Helvetica', '', 9)
                for dim, dd in mat_dims.items():
                    if isinstance(dd, dict):
                        pdf.cell(mw[0], 7, str(dim).replace('_', ' ').title(), border=1)
                        pdf.cell(mw[1], 7, f'{dd.get("score", 0):.1f}', border=1, align='C')
                        pdf.cell(mw[2], 7, str(dd.get('level', '')), border=1, align='C')
                        pdf.cell(mw[3], 7, f'{dd.get("weight", 0.2):.0%}', border=1, align='C')
                        pdf.ln()

        # ── Channel Mix Summary ──
        if section_allowed('channel_detail'):
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 18)
            pdf.cell(0, 12, 'Channel Strategy', ln=True)
            pdf.set_draw_color(255, 230, 0)
            pdf.line(10, pdf.get_y(), 80, pdf.get_y())
            pdf.ln(8)
            channel_mix = _build_channel_mix(data.get('queues', []))
            if channel_mix:
                pdf.set_font('Helvetica', 'B', 9)
                ch_headers = ['Channel', 'Volume', 'Share %', 'Avg CSAT', 'Avg AHT', 'Avg CPC']
                ch_widths = [35, 30, 25, 25, 30, 30]
                pdf.set_fill_color(46, 46, 56)
                pdf.set_text_color(255, 255, 255)
                for i, h in enumerate(ch_headers):
                    pdf.cell(ch_widths[i], 8, h, border=1, fill=True, align='C')
                pdf.ln()
                pdf.set_text_color(46, 46, 56)
                pdf.set_font('Helvetica', '', 9)
                for ch in channel_mix:
                    pdf.cell(ch_widths[0], 7, ch.get('channel', ''), border=1)
                    pdf.cell(ch_widths[1], 7, f"{ch.get('volume', 0):,}", border=1, align='C')
                    pdf.cell(ch_widths[2], 7, f"{ch.get('pct', 0):.1f}%", border=1, align='C')
                    pdf.cell(ch_widths[3], 7, f"{ch.get('avgCSAT', 0):.2f}", border=1, align='C')
                    pdf.cell(ch_widths[4], 7, f"{ch.get('avgAHT_min', ch.get('avgAHT', 0)):.0f}s", border=1, align='C')
                    pdf.cell(ch_widths[5], 7, f"${ch.get('avgCPC', 0):.2f}", border=1, align='C')
                    pdf.ln()

        # ── Diagnostic Highlights ──
        if section_allowed('diagnostic_detail'):
            pdf.ln(10)
            pdf.set_font('Helvetica', 'B', 14)
            pdf.cell(0, 10, 'Diagnostic Highlights', ln=True)
            pdf.set_font('Helvetica', '', 9)
            diag_text_items = [
                f"Total monthly volume: {data.get('totalVolume', 0):,} contacts across {len(data.get('queues', []))} queues",
                f"Business units: {', '.join(data.get('bus', []))}",
                f"Active channels: {', '.join(data.get('channels', []))}",
                f"Enabled initiatives: {sum(1 for i in STATE['initiatives'] if i.get('enabled'))} of {len(STATE['initiatives'])}",
            ]
            for item in diag_text_items:
                pdf.cell(0, 6, item, ln=True)

        # ── Channel Migration Summary ──
        if section_allowed('channel_migration'):
            migrations = STATE.get('channelStrategy', {}).get('migrations', [])
            if migrations:
                pdf.ln(10)
                pdf.set_font('Helvetica', 'B', 14)
                pdf.set_text_color(46, 46, 56)
                pdf.cell(0, 10, 'Channel Migration Summary', ln=True)
                pdf.set_font('Helvetica', 'B', 9)
                mig_widths = [50, 50, 30, 50]
                for h, w in zip(['From', 'To', 'Volume', 'Intent'], mig_widths):
                    pdf.cell(w, 7, h, border=1, align='C')
                pdf.ln()
                pdf.set_font('Helvetica', '', 9)
                for m in migrations[:15]:
                    pdf.cell(mig_widths[0], 7, str(m.get('from', '')), border=1)
                    pdf.cell(mig_widths[1], 7, str(m.get('to', '')), border=1)
                    pdf.cell(mig_widths[2], 7, f"{m.get('volume', 0):,}", border=1, align='C')
                    pdf.cell(mig_widths[3], 7, str(m.get('intent', ''))[:20], border=1)
                    pdf.ln()

        # ── Operating Model Target ──
        if section_allowed('operating_model'):
            om = STATE['overrides'].get('operating_model', {})
            if om:
                pdf.ln(10)
                pdf.set_font('Helvetica', 'B', 14)
                pdf.set_text_color(46, 46, 56)
                pdf.cell(0, 10, 'Target Operating Model', ln=True)
                pdf.set_font('Helvetica', '', 9)
                pdf.cell(0, 6, f"Onshore: {om.get('onshore', '?')}%  |  Nearshore: {om.get('nearshore', '?')}%  |  Offshore: {om.get('offshore', '?')}%", ln=True)
                tiers = om.get('tiers', [])
                for t in tiers:
                    pdf.cell(0, 6, f"  {t.get('name', '?')} — {t.get('pct', '?')}% volume share", ln=True)

        # ── Disclaimer ──
        pdf.ln(6)
        pdf.set_font('Helvetica', 'I', 8)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 6, 'Generated by ContactIQ — EY Contact Centre Opportunity Assessment', ln=True, align='C')
        pdf.cell(0, 6, f'Report scope: {role_label} | {mode.capitalize()} mode', ln=True, align='C')
        
        fd, export_path = tempfile.mkstemp(suffix='.pdf')
        os.close(fd)
        pdf.output(export_path)
        return send_file(export_path, as_attachment=True, download_name=f"ContactIQ_{params.get('clientName', 'Client')}_Assessment.pdf")
    except ImportError as e:
        return jsonify({'error': f'PDF export dependency missing: {e}. Install with: pip install fpdf2'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def _build_channel_mix(queues):
    mix = {}
    for q in queues:
        ch = q['channel']
        if ch not in mix: mix[ch] = {'channel':ch,'volume':0,'csat_w':0,'aht_w':0,'cpc_w':0}
        mix[ch]['volume'] += q['volume']; mix[ch]['csat_w'] += q.get('csat',0)*q['volume']
        mix[ch]['aht_w'] += q.get('aht',0)*q['volume']; mix[ch]['cpc_w'] += q.get('cpc',0)*q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    result = []
    for m in mix.values():
        v = m['volume']; aht_min = m['aht_w']/max(v,1)
        result.append({'channel':m['channel'],'volume':v,'pct':round(v/total*100,1),
                       'avgCSAT':round(m['csat_w']/max(v,1),2),'avgAHT':round(aht_min*60,0),
                       'avgAHT_min':round(aht_min,1),'avgCPC':round(m['cpc_w']/max(v,1),2)})
    result.sort(key=lambda x: x['volume'], reverse=True)
    return result


def _weighted_avg(queues, field):
    """Volume-weighted average of a queue-level metric."""
    total_vol = sum(q.get('volume', 0) for q in queues) or 1
    return round(sum(q.get(field, 0) * q.get('volume', 0) for q in queues) / total_vol, 4)

def _build_bu_mix(queues):
    mix = {}
    for q in queues:
        bu = q.get('bu','Unknown')
        if bu not in mix: mix[bu] = {'bu':bu,'volume':0,'csat_w':0}
        mix[bu]['volume'] += q['volume']; mix[bu]['csat_w'] += q.get('csat',0)*q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    return sorted([{'bu':m['bu'],'volume':m['volume'],'pct':round(m['volume']/total*100,1),
                    'avgCSAT':round(m['csat_w']/max(m['volume'],1),2)} for m in mix.values()],
                  key=lambda x: x['volume'], reverse=True)

def _build_intent_mix(queues):
    mix = {}
    for q in queues:
        intent = q.get('intent','Unknown')
        if intent not in mix: mix[intent] = {'intent':intent,'volume':0,'csat_w':0}
        mix[intent]['volume'] += q['volume']; mix[intent]['csat_w'] += q.get('csat',0)*q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    return sorted([{'intent':m['intent'],'volume':m['volume'],'pct':round(m['volume']/total*100,1),
                    'avgCSAT':round(m['csat_w']/max(m['volume'],1),2)} for m in mix.values()],
                  key=lambda x: x['volume'], reverse=True)

def _build_heatmap(queues):
    return [{'bu':q.get('bu',''),'intent':q.get('intent',''),'channel':q['channel'],
             'volume':q['volume'],'csat':q.get('csat',0),'aht':q.get('aht',0),
             'fcr':q.get('fcr',0),'cpc':q.get('cpc',0)} for q in queues]


def _enrich_sub_intents_for_downstream(sub_intent_analysis, initiatives, waterfall):
    """
    v12-#35: Enrich sub-intent data with initiative mapping and pool consumption.
    This data feeds into: Opportunity Buckets, Self-Service Feasibility, Channel Strategy.
    """
    enabled = [i for i in initiatives if i.get('enabled')]
    enriched = []
    for intent_data in sub_intent_analysis:
        intent = intent_data.get('intent', '')
        sub_intents = []
        for si in intent_data.get('subIntents', []):
            # Map initiatives to this sub-intent based on complexity and channel match
            complexity = si.get('complexity', 'moderate')
            auto_potential = si.get('automationPotential', 0)
            matched_initiatives = []
            for init in enabled:
                lever = init.get('lever', '')
                if complexity == 'simple' and lever in ('deflection', 'automation', 'self_service'):
                    matched_initiatives.append({'id': init['id'], 'name': init['name'], 'lever': lever})
                elif complexity == 'moderate' and lever in ('aht_reduction', 'automation', 'process_improvement'):
                    matched_initiatives.append({'id': init['id'], 'name': init['name'], 'lever': lever})
                elif complexity == 'complex' and lever in ('agent_assist', 'escalation_reduction', 'knowledge_mgmt'):
                    matched_initiatives.append({'id': init['id'], 'name': init['name'], 'lever': lever})

            # Feasibility score for self-service (used by Page 9)
            feasibility = round(auto_potential * 100)
            if complexity == 'complex': feasibility = min(feasibility, 30)
            elif complexity == 'moderate': feasibility = min(feasibility, 70)

            # Deflectable channels (used by Channel Strategy)
            deflectable_channels = []
            if auto_potential > 0.7: deflectable_channels.extend(['IVR', 'App/Self-Service', 'Chat'])
            elif auto_potential > 0.4: deflectable_channels.extend(['Chat', 'App/Self-Service'])
            elif auto_potential > 0.2: deflectable_channels.append('Chat')

            sub_intents.append({
                **si,
                'initiatives': matched_initiatives[:5],
                'feasibilityScore': feasibility,
                'deflectableChannels': deflectable_channels,
                'capturedFTE': 0,  # Will be populated by waterfall consumption
                'untappedFTE': 0,
            })

        enriched.append({
            'intent': intent,
            'totalVolume': intent_data.get('totalVolume', 0),
            'automationPotential': intent_data.get('automationPotential', 0),
            'totalDeflectable': intent_data.get('totalDeflectable', 0),
            'subIntents': sub_intents,
        })

    # CR-09: Apply sub-intent overrides from STATE
    overrides = STATE.get('overrides', {})
    for e in enriched:
        for si in e.get('subIntents', []):
            okey = f"subintent_{e['intent']}_{si.get('name', si.get('subIntent', ''))}"
            if okey in overrides:
                ovr = overrides[okey]
                if 'volShare' in ovr: si['volumeShare'] = ovr['volShare']
                if 'complexity' in ovr: si['complexity'] = ovr['complexity']
                if 'lever' in ovr: si['primaryLever'] = ovr['lever']
                if 'deflectable' in ovr:
                    si['feasibilityScore'] = 60 if ovr['deflectable'] else 10
                if 'fteOverride' in ovr: si['fteOverride'] = ovr['fteOverride']
    return enriched

def _build_cost_breakdown(data):
    total = data.get('totalCost',1) or 1
    return sorted([{'role':r['role'],'headcount':r['headcount'],'costPerFTE':r['costPerFTE'],
                    'totalCost':round(r['headcount']*r['costPerFTE']),
                    'pct':round(r['headcount']*r['costPerFTE']/total*100,1)} for r in data['roles']],
                  key=lambda x: x['totalCost'], reverse=True)


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
